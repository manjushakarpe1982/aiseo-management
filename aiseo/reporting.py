"""
Step 8 — generate_report(scan_id)

Exports an Excel workbook with 4 sheets:
  Sheet 1: Scan Summary         — totals, status breakdown, prompt versions used
  Sheet 2: Cannibalization Issues  — full detail + status + reasoning + prompt version
  Sheet 3: Content Improvements    — current vs suggested + char counts + reasoning
  Sheet 4: Page Keywords           — primary keyword, intent, gaps, LSI per page
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .db import get_connection
from .config import TP


# ── Style constants ───────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_TITLE_FONT    = Font(bold=True, name="Calibri", size=14)
_SUBHEAD_FONT  = Font(bold=True, name="Calibri", size=11)
_BODY_FONT     = Font(name="Calibri", size=10)
_WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGN     = Alignment(vertical="top")

_THIN = Side(style="thin", color="CCCCCC")
_CELL_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_SEVERITY_FILLS = {
    "High":   PatternFill("solid", fgColor="FFCCCC"),
    "Medium": PatternFill("solid", fgColor="FFF2CC"),
    "Low":    PatternFill("solid", fgColor="E2EFDA"),
}
_PRIORITY_FILLS = _SEVERITY_FILLS  # same palette

_STATUS_FILLS = {
    "Yet to Act": PatternFill("solid", fgColor="DEEBF7"),
    "Acted":      PatternFill("solid", fgColor="E2EFDA"),
    "Deferred":   PatternFill("solid", fgColor="FFF2CC"),
}


def _style_header_row(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def _set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _body_cell(ws, row: int, col: int, value, wrap: bool = False,
               fill=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = _BODY_FONT
    cell.border = _CELL_BORDER
    cell.alignment = _WRAP_ALIGN if wrap else _TOP_ALIGN
    if fill:
        cell.fill = fill


# ── Sheet 1: Scan Summary ─────────────────────────────────────────────────

def _build_summary_sheet(ws, conn, scan_id: int) -> None:
    ws.title = "Scan Summary"

    cursor = conn.cursor()

    # Scan metadata
    cursor.execute(
        f"""
        SELECT s.ScanID, s.RunID, s.ScanName, s.StartedAt, s.EndedAt,
               s.TotalURLs, s.URLsScraped, s.TreesAnalysed, s.Status,
               cp.VersionLabel AS CannibalizationPromptVersion,
               pp.VersionLabel AS ContentPromptVersion,
               u.FullName AS StartedBy
        FROM {TP}Scans s
        LEFT JOIN {TP}Prompts cp ON s.CannibalizationPromptID = cp.PromptID
        LEFT JOIN {TP}Prompts pp ON s.ContentPromptID         = pp.PromptID
        LEFT JOIN {TP}Users   u  ON s.StartedByUserID         = u.UserID
        WHERE s.ScanID = ?
        """,
        (scan_id,),
    )
    scan = cursor.fetchone()
    if not scan:
        ws["A1"] = f"Scan ID {scan_id} not found."
        return

    # Title
    ws["A1"] = "AISEO Scan Summary Report"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 24

    # Metadata block
    meta = [
        ("Scan ID",                   scan[0]),
        ("Run ID",                    scan[1]),
        ("Scan Name",                 scan[2]),
        ("Status",                    scan[8]),
        ("Started At",                scan[3].strftime("%Y-%m-%d %H:%M") if scan[3] else ""),
        ("Ended At",                  scan[4].strftime("%Y-%m-%d %H:%M") if scan[4] else "-"),
        ("Started By",                scan[11]),
        ("Total URLs",                scan[5]),
        ("URLs Scraped",              scan[6]),
        ("Trees Analysed",            scan[7]),
        ("Cannibalization Prompt",    scan[9]),
        ("Content Prompt",            scan[10]),
    ]
    for r, (label, value) in enumerate(meta, start=3):
        cell_lbl = ws.cell(row=r, column=1, value=label)
        cell_lbl.font = _SUBHEAD_FONT
        cell_lbl.border = _CELL_BORDER
        cell_val = ws.cell(row=r, column=2, value=value)
        cell_val.font = _BODY_FONT
        cell_val.border = _CELL_BORDER

    # Status breakdown — Cannibalization
    row = 17
    ws.cell(row=row, column=1, value="Cannibalization Issues by Status").font = _SUBHEAD_FONT
    row += 1
    _style_header_row(ws, row, 2)
    ws.cell(row=row, column=1, value="Status")
    ws.cell(row=row, column=2, value="Count")
    row += 1
    cursor.execute(
        f"""
        SELECT Status, COUNT(*) FROM {TP}CannibalizationIssues
        WHERE ScanID = ? GROUP BY Status
        """,
        (scan_id,),
    )
    for r in cursor.fetchall():
        _body_cell(ws, row, 1, r[0], fill=_STATUS_FILLS.get(r[0]))
        _body_cell(ws, row, 2, r[1])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Content Improvements by Status").font = _SUBHEAD_FONT
    row += 1
    _style_header_row(ws, row, 2)
    ws.cell(row=row, column=1, value="Status")
    ws.cell(row=row, column=2, value="Count")
    row += 1
    cursor.execute(
        f"""
        SELECT Status, COUNT(*) FROM {TP}ContentImprovements
        WHERE ScanID = ? GROUP BY Status
        """,
        (scan_id,),
    )
    for r in cursor.fetchall():
        _body_cell(ws, row, 1, r[0], fill=_STATUS_FILLS.get(r[0]))
        _body_cell(ws, row, 2, r[1])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Content Improvements by Priority").font = _SUBHEAD_FONT
    row += 1
    _style_header_row(ws, row, 2)
    ws.cell(row=row, column=1, value="Priority")
    ws.cell(row=row, column=2, value="Count")
    row += 1
    cursor.execute(
        f"""
        SELECT Priority, COUNT(*) FROM {TP}ContentImprovements
        WHERE ScanID = ? GROUP BY Priority
        ORDER BY
            CASE Priority WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END
        """,
        (scan_id,),
    )
    for r in cursor.fetchall():
        _body_cell(ws, row, 1, r[0], fill=_PRIORITY_FILLS.get(r[0]))
        _body_cell(ws, row, 2, r[1])
        row += 1

    _set_col_widths(ws, [28, 40])


# ── Sheet 2: Cannibalization Issues ──────────────────────────────────────

def _build_cannibal_sheet(ws, conn, scan_id: int) -> None:
    ws.title = "Cannibalization Issues"

    cursor = conn.cursor()
    cursor.execute(
        f"""
        SELECT ci.IssueID, ci.TreeCluster, ci.CannibalKeyword, ci.Severity,
               ci.SeverityReason,
               ci.URL1, ci.URL1_FieldName, ci.URL1_CurrentContent, ci.URL1_SuggestedFix,
               ci.URL2, ci.URL2_FieldName, ci.URL2_CurrentContent, ci.URL2_SuggestedFix,
               ci.OverallRecommendation, ci.Reasoning,
               ci.Status, ci.UserComment, ci.DeferredReason,
               ci.VerifiedFixed, ci.CreatedAt,
               p.VersionLabel AS PromptVersion,
               u.FullName AS AuditedBy, ci.LastAuditedAt
        FROM {TP}CannibalizationIssues ci
        LEFT JOIN {TP}Prompts p ON ci.PromptID = p.PromptID
        LEFT JOIN {TP}Users   u ON ci.LastAuditedByUserID = u.UserID
        WHERE ci.ScanID = ?
        ORDER BY
            CASE ci.Severity WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
            ci.TreeCluster
        """,
        (scan_id,),
    )
    rows = cursor.fetchall()

    headers = [
        "Issue ID", "Tree", "Cannibal Keyword", "Severity", "Severity Reason",
        "URL 1", "URL1 Field", "URL1 Current Content", "URL1 Suggested Fix",
        "URL 2", "URL2 Field", "URL2 Current Content", "URL2 Suggested Fix",
        "Overall Recommendation", "Reasoning",
        "Status", "User Comment", "Deferred Reason",
        "Verified Fixed", "Created At", "Prompt Version",
        "Audited By", "Last Audited At",
    ]
    _style_header_row(ws, 1, len(headers))
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.freeze_panes = "A2"

    widths = [8,18,22,10,30,50,14,35,35,50,14,35,35,30,50,14,25,25,12,18,30,20,18]

    for r_idx, row in enumerate(rows, start=2):
        severity = row[3]
        sev_fill = _SEVERITY_FILLS.get(severity)
        status   = row[15]
        st_fill  = _STATUS_FILLS.get(status)

        data = list(row)
        data[18] = "Yes" if data[18] else "No"
        data[19] = data[19].strftime("%Y-%m-%d") if data[19] else ""
        data[22] = data[22].strftime("%Y-%m-%d") if data[22] else ""

        for col_idx, value in enumerate(data, start=1):
            wrap = col_idx in (8, 9, 12, 13, 14, 15, 17, 18)
            fill = sev_fill if col_idx == 4 else (st_fill if col_idx == 16 else None)
            _body_cell(ws, r_idx, col_idx, value, wrap=wrap, fill=fill)
        ws.row_dimensions[r_idx].height = 60

    _set_col_widths(ws, widths)


# ── Sheet 3: Content Improvements ─────────────────────────────────────────

def _build_improvements_sheet(ws, conn, scan_id: int) -> None:
    ws.title = "Content Improvements"

    cursor = conn.cursor()
    cursor.execute(
        f"""
        SELECT ci.ImprovementID, ci.TreeCluster, ci.PageURL,
               ci.FieldName, ci.IssueType, ci.Priority,
               ci.CurrentContent,  ci.CurrentCharCount,
               ci.SuggestedContent, ci.SuggestedCharCount,
               ci.Reasoning, ci.ImpactEstimate,
               ci.Status, ci.UserComment, ci.DeferredReason,
               ci.VerifiedFixed, ci.CreatedAt,
               p.VersionLabel AS PromptVersion,
               u.FullName AS AuditedBy, ci.LastAuditedAt
        FROM {TP}ContentImprovements ci
        LEFT JOIN {TP}Prompts p ON ci.PromptID = p.PromptID
        LEFT JOIN {TP}Users   u ON ci.LastAuditedByUserID = u.UserID
        WHERE ci.ScanID = ?
        ORDER BY
            CASE ci.Priority WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
            ci.PageURL, ci.FieldName
        """,
        (scan_id,),
    )
    rows = cursor.fetchall()

    headers = [
        "ID", "Tree", "Page URL",
        "Field", "Issue Type", "Priority",
        "Current Content", "Current Chars",
        "Suggested Content", "Suggested Chars",
        "Reasoning", "Impact Estimate",
        "Status", "User Comment", "Deferred Reason",
        "Verified Fixed", "Created At",
        "Prompt Version", "Audited By", "Last Audited At",
    ]
    _style_header_row(ws, 1, len(headers))
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.freeze_panes = "A2"

    widths = [8,18,50,16,22,10,40,12,40,14,50,20,14,25,25,12,18,30,20,18]

    for r_idx, row in enumerate(rows, start=2):
        priority = row[5]
        pri_fill = _PRIORITY_FILLS.get(priority)
        status   = row[12]
        st_fill  = _STATUS_FILLS.get(status)

        data = list(row)
        data[15] = "Yes" if data[15] else "No"
        data[16] = data[16].strftime("%Y-%m-%d") if data[16] else ""
        data[19] = data[19].strftime("%Y-%m-%d") if data[19] else ""

        for col_idx, value in enumerate(data, start=1):
            wrap = col_idx in (7, 9, 11, 14, 15)
            fill = pri_fill if col_idx == 6 else (st_fill if col_idx == 13 else None)
            _body_cell(ws, r_idx, col_idx, value, wrap=wrap, fill=fill)
        ws.row_dimensions[r_idx].height = 60

    _set_col_widths(ws, widths)


# ── Sheet 4: Page Keywords ────────────────────────────────────────────────

def _build_keywords_sheet(ws, conn, scan_id: int) -> None:
    ws.title = "Page Keywords"

    cursor = conn.cursor()
    cursor.execute(
        f"""
        SELECT pk.KeywordID, pk.TreeCluster, pk.PageURL,
               pk.PrimaryKeyword, pk.SecondaryKeywords, pk.SearchIntent,
               pk.KeywordGaps, pk.MissingLSITerms, pk.ContentFocusScore,
               pk.CreatedAt, p.VersionLabel AS PromptVersion
        FROM {TP}PageKeywords pk
        LEFT JOIN {TP}Prompts p ON pk.PromptID = p.PromptID
        WHERE pk.ScanID = ?
        ORDER BY pk.TreeCluster, pk.PageURL
        """,
        (scan_id,),
    )
    rows = cursor.fetchall()

    headers = [
        "ID", "Tree", "Page URL",
        "Primary Keyword", "Secondary Keywords", "Search Intent",
        "Keyword Gaps", "Missing LSI Terms", "Content Focus Score",
        "Created At", "Prompt Version",
    ]
    _style_header_row(ws, 1, len(headers))
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.freeze_panes = "A2"

    widths = [8, 18, 52, 30, 42, 16, 42, 42, 16, 18, 28]

    _INTENT_FILL = PatternFill("solid", fgColor="E2EFDA")  # green for transactional
    _FOCUS_FILLS = {
        range(8, 11): PatternFill("solid", fgColor="E2EFDA"),   # 8-10 green
        range(5, 8):  PatternFill("solid", fgColor="FFF2CC"),   # 5-7 yellow
        range(0, 5):  PatternFill("solid", fgColor="FFCCCC"),   # 0-4 red
    }

    def _focus_fill(score):
        if score is None:
            return None
        for r, fill in _FOCUS_FILLS.items():
            if score in r:
                return fill
        return None

    for r_idx, row in enumerate(rows, start=2):
        data      = list(row)
        data[9]   = data[9].strftime("%Y-%m-%d") if data[9] else ""

        focus_score = data[8]

        for col_idx, value in enumerate(data, start=1):
            wrap = col_idx in (3, 4, 5, 7, 8)
            fill = None
            if col_idx == 6 and value == "transactional":
                fill = _INTENT_FILL
            elif col_idx == 9:
                fill = _focus_fill(focus_score)
            _body_cell(ws, r_idx, col_idx, value, wrap=wrap, fill=fill)
        ws.row_dimensions[r_idx].height = 50

    _set_col_widths(ws, widths)


# ── Public entry point ────────────────────────────────────────────────────

def generate_report(scan_id: int, output_dir: str = ".") -> str:
    """
    Build the Excel report for the given ScanID.

    Returns the file path of the generated .xlsx file.
    """
    conn = get_connection()

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Scan Summary")
    ws2 = wb.create_sheet("Cannibalization Issues")
    ws3 = wb.create_sheet("Content Improvements")
    ws4 = wb.create_sheet("Page Keywords")

    _build_summary_sheet(ws1, conn, scan_id)
    _build_cannibal_sheet(ws2, conn, scan_id)
    _build_improvements_sheet(ws3, conn, scan_id)
    _build_keywords_sheet(ws4, conn, scan_id)

    conn.close()

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"AISEO_Report_Scan{scan_id}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"Report saved: {filepath}")
    return filepath
